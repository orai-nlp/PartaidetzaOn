from partaidetza.partaidetza_app.forms import LoginForm, SearchBoxForm
from partaidetza.partaidetza_app.models import Proposal
from django.contrib.auth import authenticate, login as auth_login
from partaidetza.settings import LANGUAGES, MEDIA_ROOT
import os, urllib2
from openpyxl import Workbook, load_workbook
from partaidetza.settings import LANGUAGES, OPENTRAD_CODE, AT_LANGUAGE_PRIORITY, \
TEXT_CAT_PATH, TEXT_CAT_MODELS_PATH, TEXT_CAT_LANGUAGE_CONVERTOR

def log_in(request):
    """Logs a user who has entered your NAN and password
        PARAMETERS:
        1. request element
    """            
    def _log_in(NAN,password):
        """Logs a user who has entered your NAN and password
        PARAMETERS:
        1. NAN: User's NAN
        2. password: User's password"""
        user = authenticate(username = NAN, password = password)     
                          
        # return User element  
        return user
       
       
    login_form = LoginForm(request.POST)
    
    # Login form is valid
    if login_form.is_valid():
        cd = login_form.cleaned_data
        
        user = _log_in(cd.get("NAN"),cd.get("password"))    
        if user is not None:
            if user.is_active:
                auth_login(request, user)

                return user.profile
        else:
            return None
    else: # Login form is not valid
        return None
                
 
    
def base_forms_initialization(request):
    """Initializes various common forms templates
    PARAMETERS:
    1. request: request element
    """ 
    login_form = LoginForm()
    search_form = SearchBoxForm(request.POST)
    
    return (login_form, search_form)
    
    
def validate_new_event_form(event_dict, event_date_form):
    """Validates all form included in add_event template"""
    validated = event_date_form.is_valid()
    for lang in LANGUAGES:
        validated = validated and event_dict[lang[0]].is_valid()
    return validated
   
    
def validate_new_proposal_form(new_proposal_form, title_dict, summary_dict, necessity_dict, recipient_dict, where_dict, upload_form, coordinates_form):
    """Validates all form included in add_proposal template"""
    validated = new_proposal_form.is_valid()
    for lang in LANGUAGES:
        validated = validated and \
                    title_dict[lang[0]].is_valid() and \
                    summary_dict[lang[0]].is_valid() and \
                    necessity_dict[lang[0]].is_valid() and \
                    recipient_dict[lang[0]].is_valid() and \
                    where_dict[lang[0]].is_valid() and \
                    upload_form.is_valid() and \
                    coordinates_form.is_valid()
    return validated

def validate_new_accepted_proposal_form(new_proposal_form, title_dict, summary_dict, necessity_dict, recipient_dict, where_dict, source_dict, explanation_dict, upload_form, coordinates_form):
    """Validates all form included in add_accepted_proposal template"""
    validated = new_proposal_form.is_valid()
    for lang in LANGUAGES:
        validated = validated and \
                    title_dict[lang[0]].is_valid() and \
                    summary_dict[lang[0]].is_valid() and \
                    necessity_dict[lang[0]].is_valid() and \
                    recipient_dict[lang[0]].is_valid() and \
                    where_dict[lang[0]].is_valid() and \
                    source_dict[lang[0]].is_valid() and \
                    explanation_dict[lang[0]].is_valid() and \
                    upload_form.is_valid() and\
                    coordinates_form.is_valid()
    return validated

def handle_uploaded_file(name,f,type='upload'):
    try:
        if type == 'upload': # standard file upload
            path = MEDIA_ROOT
        elif type == 'xls': # xls file upload for management actions
            path = MEDIA_ROOT + "/xls"
        if not os.path.exists("/".join((path+"/"+name).split('/')[:-1])):
            print "/".join((path+"/"+name).split('/')[:-1])
            os.makedirs("/".join((path+"/"+name).split('/')[:-1]))
        destination = open(path+"/"+name, 'wb+')
        for chunk in f.chunks():
            destination.write(chunk)
        destination.close()
    except Exception as a:
        print a
        return False
    return True

def login_needed(request):
    """Returns true in login box needs to be visible"""
    return "?next=" in request.build_absolute_uri()
    
  

  
def extract_project_info_from_file(filename, type='xls'):
    """Extracs project info information from a file"""
    status_code =1
    # FILE FORMAT: TODO
    
    def None_to_str(x):
        if x is None:
            return "" 
        else:
            return x 
    try:
        if type == 'xls':
            path = MEDIA_ROOT + "/xls"
            # load the Worbook class from file
            wb = load_workbook(path+'/'+filename)
            # get the first worksheet
            work_sheet = wb.worksheets[0]
            # get all rows
            rows = work_sheet.rows
            # get metadata rows
            metadata_row = rows[0]
            # get data rows
            data_rows = rows[1:]            
            rows = []
            # for each data row
            for data_row in data_rows:
                # for each metadata column save the value
                row = {}
                for metadata_column in metadata_row:                
                    row[metadata_column.value] = None_to_str(data_row[(metadata_row.index(metadata_column))].value)
                rows+=[row]
            return (status_code, rows)
    except Exception as error:
        status_code = 0
        print error
        return (status_code, error)        
    
 
def create_project_info_file(proposal_id):
    """Creates a project info XML that can be used to create a new project"""
    proposal = Proposal.objects.get(id=int(proposal_id))
    path = MEDIA_ROOT + "/xls"
    wb = wb = Workbook()
    ws = wb.create_sheet(0)
    # METADATA ROW
    ws.cell(row = 1, column = 1).value = "id"
    ws.cell(row = 1, column = 2).value = "cost"
    ws.cell(row = 1, column = 3).value = "recipient_number"
    ws.cell(row = 1, column = 4).value = "area"
    # LANGUAGE COLUMNS
    column_index = 5
    for lang in LANGUAGES:
        ws.cell(row = 1, column = column_index).value = "title_"+lang[0]
        column_index += 1
        ws.cell(row = 1, column = column_index).value = "summary_"+lang[0]
        column_index += 1
        ws.cell(row = 1, column = column_index).value = "necessity_"+lang[0]
        column_index += 1
        ws.cell(row = 1, column = column_index).value = "recipient_"+lang[0]
        column_index += 1
        ws.cell(row = 1, column = column_index).value = "where_"+lang[0]
        column_index += 1
        ws.cell(row = 1, column = column_index).value = "source_"+lang[0]
        column_index += 1
        ws.cell(row = 1, column = column_index).value = "explanation_"+lang[0]
    # DATA ROW
    ws.cell(row = 2, column = 1).value = proposal.id
    ws.cell(row = 2, column = 2).value = proposal.cost
    ws.cell(row = 2, column = 3).value = proposal.recipient_number
    ws.cell(row = 2, column = 4).value = ",".join(map(lambda x: str(x),proposal.get_area()))
    column_index = 5
    for lang in LANGUAGES:
        ws.cell(row = 2, column = column_index).value = proposal.get_title(lang[0])
        column_index += 1
        ws.cell(row = 2, column = column_index).value = proposal.get_summary(lang[0])
        column_index += 1
        ws.cell(row = 2, column = column_index).value = proposal.get_necessity(lang[0])
        column_index += 1
        ws.cell(row = 2, column = column_index).value = proposal.get_recipient(lang[0])
        column_index += 1
        ws.cell(row = 2, column = column_index).value = proposal.get_where(lang[0])
        column_index += 1
        ws.cell(row = 2, column = column_index).value = ""
        column_index += 1
        ws.cell(row = 2, column = column_index).value = ""
    # CREATE FILE
    wb.save(path+'/proposal_file_'+proposal_id+'.xlsx')
    return path+'/proposal_file_'+proposal_id+'.xlsx'
    

def create_multiple_project_info_file(proposal_id_list):
    """Creates a project info XML that can be used to create a new projects"""
    path = MEDIA_ROOT + "/xls"
    wb = wb = Workbook()
    ws = wb.create_sheet(0)
    # METADATA ROW
    ws.cell(row = 1, column = 1).value = "id"
    ws.cell(row = 1, column = 2).value = "cost"
    ws.cell(row = 1, column = 3).value = "recipient_number"
    ws.cell(row = 1, column = 4).value = "area"
    # LANGUAGE COLUMNS
    column_index = 5
    for lang in LANGUAGES:
        ws.cell(row = 1, column = column_index).value = "title_"+lang[0]
        column_index += 1
        ws.cell(row = 1, column = column_index).value = "summary_"+lang[0]
        column_index += 1
        ws.cell(row = 1, column = column_index).value = "necessity_"+lang[0]
        column_index += 1
        ws.cell(row = 1, column = column_index).value = "recipient_"+lang[0]
        column_index += 1
        ws.cell(row = 1, column = column_index).value = "where_"+lang[0]
    row_index=2
    for proposal_id in proposal_id_list:
        proposal = Proposal.objects.get(id=int(proposal_id))
        # DATA ROW
        ws.cell(row = row_index, column = 1).value = proposal.id
        ws.cell(row = row_index, column = 2).value = proposal.cost
        ws.cell(row = row_index, column = 3).value = proposal.recipient_number
        ws.cell(row = row_index, column = 4).value = ",".join(map(lambda x: str(x),proposal.get_area()))
        column_index = 5
        for lang in LANGUAGES:
            ws.cell(row = row_index, column = column_index).value = proposal.get_title(lang[0])
            column_index += 1
            ws.cell(row = row_index, column = column_index).value = proposal.get_summary(lang[0])
            column_index += 1
            ws.cell(row = row_index, column = column_index).value = proposal.get_necessity(lang[0])
            column_index += 1
            ws.cell(row = row_index, column = column_index).value = proposal.get_recipient(lang[0])
            column_index += 1
            ws.cell(row = row_index, column = column_index).value = proposal.get_where(lang[0])
        row_index +=1
        # CREATE FILE
    id_path = "_".join(proposal_id_list)
    wb.save(path+'/proposal_file_'+id_path+'.xlsx')
    return path+'/proposal_file_'+id_path+'.xlsx'
    
        
# AUTOMATIC_TRANSLATION FUNCTIONS

def get_best_translation_source(dict,target_lang):
    """Get the best translation source"""
    translated_text = ""
    for source_lang in AT_LANGUAGE_PRIORITY:
        if source_lang != target_lang:
            if dict[source_lang] != '':
                translated_text = translate_text_opentrad(dict[source_lang].cleaned_data.get(\
                                                        dict[source_lang].cleaned_data.keys()[0]), source_lang, target_lang)
                if "Error: Mode " not in translated_text:
                    break
    return translated_text

def translate_text_opentrad(text, source_language, target_language):
    """Translate text automatically from source_language to target_language"""    
    
    url="http://api.opentrad.com/translate.php"
    url_text="?text="+text.strip()
    url_lang="&lang={0}-{1}".format(source_language,target_language)
    url_client="&cod_client="+OPENTRAD_CODE
    URL=url+url_text+url_lang+url_client

    try:
        return urllib2.urlopen(URL.replace(' ','%20').encode("utf-8")).read()
    except:
        return urllib2.urlopen(URL.replace(' ','%20')).read()    
        
    
# LANGUAGE DETECTION FUNCTIONS    
        
def get_text_language(text):
    """Returns the language(s) of the text in a list"""
    from subprocess import check_output
    lang = check_output([TEXT_CAT_PATH+"/text_cat", "-d", TEXT_CAT_MODELS_PATH, "-l", text])
    print "TEXTCAT: ",lang
    languages = lang.strip().split('or')
    languages = map(lambda x: TEXT_CAT_LANGUAGE_CONVERTOR.get(x.strip(),""),languages)
    return languages
    
def correct_language(text,lang):
    """Returns if the expected language is OK"""
    print lang, get_text_language(text)
    return lang in get_text_language(text)
